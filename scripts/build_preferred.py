#!/usr/bin/env python3
"""Convert 通用料明细.xlsx to a simple JSON lookup for the server."""
import openpyxl, json, os

SRC = os.path.join(os.path.dirname(__file__), '..', 'raw', '通用料明细.xlsx')
DST = os.path.join(os.path.dirname(__file__), '..', 'web', 'public', 'data', 'preferred_pns.json')

wb = openpyxl.load_workbook(SRC)
ws = wb.active
pns = {}
for r in range(2, ws.max_row + 1):
    brand = str(ws.cell(r, 2).value or '').strip()
    pn = str(ws.cell(r, 3).value or '').strip()
    if pn:
        pns[pn.upper()] = brand

with open(DST, 'w') as f:
    json.dump(pns, f, ensure_ascii=False)
print(f'Wrote {len(pns)} preferred PNs to {DST}')
